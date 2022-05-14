import datetime
import json
import re
import threading
import time
import math

import requests
from lxml import etree
import openpyxl

class ZhenfucaigouSpider():
    url = 'http://search.ccgp.gov.cn/bxsearch?searchtype=2'
    keyword = '福建师范大学'
    start_time = '2020:01:01'
    end_time = '2020:10:09'
    page_num = 1
    Tag =2

    params = {
        'searchtype': '2',
        'page_index': page_num,
        'bidSort': '0',
        'pinMu': '0',
        'bidType': '7',
        'kw': keyword,
        'start_time': start_time,
        'end_time': end_time,
        'timeType': '6'
    }
    headers = {
        'Cookie': 'JSESSIONID=EgPd86-6id_etA2QDV31Kks3FrNs-4gwHMoSmEZvnEktWIakHbV3!354619916; Hm_lvt_9f8bda7a6bb3d1d7a9c7196bfed609b5=1602214804; Hm_lpvt_9f8bda7a6bb3d1d7a9c7196bfed609b5=1602214892; JSESSIONID=OBoLczbR_k89lC8sOuKF4W-46DVqKEd5u7isUpSyOjE6D0nBP94c!1675672049; Hm_lvt_9459d8c503dd3c37b526898ff5aacadd=1602214902,1602214928,1602214932,1602214937; Hm_lpvt_9459d8c503dd3c37b526898ff5aacadd=1602214937',
        'Host': 'search.ccgp.gov.cn',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.121 Safari/537.36'
    }

    def get_page(self,url,headers,params):
        try:
            response = requests.get(url,headers=headers,params=params)
            if response.status_code == 200:
                html = response.content.decode('utf-8', 'ignore').replace(u'xa9', u'')
                return html
            else:
                print('status_code : '+ str(response.status_code))
                print(response.content.decode('utf-8', 'ignore').replace(u'xa9', u''))
                return None
        except requests.ConnectionError:
            return None

    def get_detail_page(self,url):
        try:
            response = requests.get(url)
            if response.status_code == 200:
                html = response.content.decode('utf-8', 'ignore').replace(u'xa9', u'')
                #print(html)
                return html
        except requests.ConnectionError:
            return None

    def get_all_url(self,html):
        pattern1 = '<.*?(href=".*?htm").*?'
        href_url = re.findall(pattern1, html, re.I)
        #print(href_url)
        url_list = []

        for url in href_url:
            url1 = url.replace('href=','').replace('"','')
            url_list.append(url1)
            #table.cell(row=i, column=2).value = url1
        #print("url_list=",url_list)
        return url_list

    # def parse_datail_page(self,html):
    #     table_list = html.xpath('//div[@class="table"]//tr')
    #     print("table_list",table_list)
    #     all_info = {}
    #     for table in table_list:
    #         if len(table.xpath('td[@class="title"]/text()'))>0:
    #             #print(''.join(table.xpath('td[@class="title"]/text()'))+":"+''.join(table.xpath('td[@colspan="3"]/text()')))
    #             title = ''.join(table.xpath('td[@class="title"]/text()'))
    #             value = ''.join(table.xpath('td[@colspan="3"]/text()'))
    #             if (title.find('附件')==0):
    #                 value = 'http://www.ccgp.gov.cn/oss/download?uuid='+''.join(table.xpath('td[@colspan="3"]/a/@id'))
    #                 #print(title+value)
    #             if ('公告时间' in title):
    #                 title = '公告时间'
    #                 value = table.xpath('td[@width="168"]/text()')[1]
    #                 district_key = '行政区域'
    #                 district_value = (table.xpath('td[@width="168"]/text()'))[0]
    #                 all_info[district_key]=district_value
    #             if '本项目招标公告日期中标日期' in title :
    #                 title = '本项目招标公告日期'
    #                 value = table.xpath('td[@width="168"]/text()')[0]
    #                 zhongbiaoriqi_key = '中标日期'
    #                 zhongbiaoriqi_value = table.xpath('td[@width="168"]/text()')[1]
    #                 all_info[zhongbiaoriqi_key]=zhongbiaoriqi_value
    #                 #print('中标日期'+zhongbiaoriqi_value)
    #             if '本项目招标公告日期成交日期' in title:
    #                 title = '本项目招标公告日期'
    #                 value = table.xpath('td[@width="168"]/text()')[0]
    #                 zhongbiaoriqi_key = '中标日期'
    #                 zhongbiaoriqi_value = ''.join(table.xpath('td[@width="168"]/text()'))[11:]
    #                 #print('zhongbiaoriqi_value:'+zhongbiaoriqi_value)
    #                 all_info[zhongbiaoriqi_key] = zhongbiaoriqi_value
    #             all_info[title] = value
    #             all_info['插入时间']= datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    #     return all_info
    #     #return json.dumps(all_info,ensure_ascii=False)
    def parse_datail_page(self,html,url,table):
        title= html.xpath("//title//text()")
        # print("title==",title)
        # print("url==",url)
        #table.cell(row=i, column=1).value = title
        print(self.Tag)
        table.cell(row=self.Tag, column=1).value = str(title)
        table.cell(row=self.Tag, column=2).value = url
        self.Tag +=1

    def start(self,url,table):
        time.sleep(0.01)
        # print(url)
        html = self.get_detail_page(url)
        html = etree.HTML(html)
        print("html2=", html)
        all_info = self.parse_datail_page(html,url,table)
        #print(all_info)
        #print(all_info.keys())

    def pages_num(self,html):
        num_list=html.xpath('/html/body/div[5]/div[1]/div/p/span[2]/text()')
        num = int(num_list[0])  #转换 int型
        # print(num)
        return num


    def run(self):
        execl_path = "D:\我的资料\python\ccgp_gov_crawler//123.xlsx" #要先在D盘创建2020年.xlsx文件
        wb = openpyxl.load_workbook(execl_path)
        table = wb['Sheet1']
        #获取搜索检索之后共有多少条内容
        html1 = self.get_page(url=self.url, headers=self.headers, params=self.params)
        html1 = etree.HTML(html1)
        pagesNum = self.pages_num(html1)
        #向上取整数
        pagesNum = math.ceil(pagesNum/20)
        print("pagesNum==",pagesNum)

        for i in range(1,pagesNum+1):
            print('正在爬取第{}页'.format(str(i)))
            self.params['page_index']=i
            html = self.get_page(url=self.url, headers=self.headers, params=self.params)
            #print(html)
            if(None == html):
                return None
            url_list = self.get_all_url(html)

            # 创建线程
            threads = []
            files = range(len(url_list))


            for url in url_list:
                t = threading.Thread(target=self.start(url,table), args=url)
                threads.append(t)

            # 启动线程
            for i in files:
                threads[i].start()
            for i in files:
                threads[i].join()
        wb.save(execl_path)


if __name__ == '__main__':
    zhenfucaigouSpider = ZhenfucaigouSpider()
    zhenfucaigouSpider.run()